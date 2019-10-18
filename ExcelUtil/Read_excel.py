from openpyxl import load_workbook
class ReadExcel(object):
    def __init__(self,excelname,excelsheet):
        self.wb=load_workbook(excelname)
        self.sheet=self.wb.get_sheet_by_name(excelsheet)
        self.maxRownum=self.sheet.max_row#获取最大行数
        self.maxColumn=self.sheet.max_column#获取最大列数


    def getValue(self):
        datalist=[]
        for row in self.sheet.rows[0:self.maxRownum]:
            list=[]
            for column in range(self.maxColumn):
                list.append(row[column].value)
            datalist.append(list)
        print(datalist)


if __name__=="__main__":
    value=ReadExcel(r"G:\Myproject\ExcelUtil\test2.xlsx","Sheet1")
    value.getValue()
