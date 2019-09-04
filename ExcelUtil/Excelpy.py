from openpyxl import load_workbook
class ParaExcel(object):
    def __init__(self,excelpath,sheetname):
        self.wb=load_workbook(excelpath)
        self.sheet=self.wb.get_sheet_by_name(sheetname)
        self.maxRownum=self.sheet.max_row#当前sheet的最大行数
        self.raxColumn=self.sheet.max_column#当前sheet的最大列数

    def getData(self):
        datalist=[]
        for line in self.sheet.rows[1:self.maxRownum]:
            list=[]
            list.append(line[1].value)
            list.append(line[2].value)
            #list.append(line[3].value)
            datalist.append(list)
        return datalist
if __name__=="__main__":
    getdata=ParaExcel(r"E:\Users\Administrator\PycharmProjects\untitled\repuests\File\casedata\AccountLoginData.xlsx","Sheet1")
    print(getdata.getData())

